"""
Constructor de libros Excel con openpyxl. Lógica pura (recibe dicts de resultados
producidos por los solvers y arma un .xlsx con formato educativo).
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAL = "0F766E"
TEAL_L = "D5F0EC"
PURPLE = "6D28D9"
GREY = "F1F5F9"
GREEN = "DCFCE7"

_thin = Side(style="thin", color="CBD5E1")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
H_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color=TEAL, size=15)
SUB_FONT = Font(italic=True, color="64748B", size=9)
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _title(ws, text, sub=None, row=1):
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    if sub:
        ws.cell(row=row + 1, column=1, value=sub).font = SUB_FONT
    return row + (3 if sub else 2)


def _header_row(ws, headers, row, fill=TEAL):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = H_FONT
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = CENTER
        c.border = BORDER
    return row + 1


def _autosize(ws, maxw=42):
    for col in ws.columns:
        length = 0
        letter = None
        for cell in col:
            if letter is None:
                letter = cell.column_letter
            v = cell.value
            if v is not None:
                length = max(length, len(str(v)))
        if letter:
            ws.column_dimensions[letter].width = min(maxw, max(10, length + 2))


def _num(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.border = BORDER
    if isinstance(v, (int, float)):
        cell.number_format = "#,##0.0000"
        cell.alignment = Alignment(horizontal="right")
    return cell


def _save(wb):
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def _footer(ws, row):
    ws.cell(row=row + 1, column=1,
            value=f"PronoStat · generado {datetime.now():%Y-%m-%d %H:%M} · Dr. César Ortiz Méndez · USACH").font = SUB_FONT


def build_informe(res, source="dataset"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe"
    r = _title(ws, "PronoStat — Informe automático del conjunto de datos",
               f"Fuente: {source} · {res['n']} filas · {res['n_cols']} columnas ({res['n_num']} numéricas)")
    ws.cell(row=r, column=1, value="Hallazgos").font = BOLD; r += 1
    for ins in res["insights"]:
        c = ws.cell(row=r, column=1, value="•  " + ins)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Perfil por columna").font = BOLD; r += 1
    r = _header_row(ws, ["Columna", "Tipo", "Faltantes %", "Media", "Mediana", "Desv.", "CV%", "Mín", "Máx", "Asimetría", "Atípicos"], r)
    for c in res["columnas"]:
        ws.cell(row=r, column=1, value=c["name"]).border = BORDER
        ws.cell(row=r, column=2, value=c["type"]).border = BORDER
        _num(ws, r, 3, c["faltantes_pct"])
        s = c.get("stats")
        if s:
            for j, k in [(4, "media"), (5, "mediana"), (6, "sd"), (7, "cv"), (8, "min"), (9, "max"), (10, "asimetria")]:
                _num(ws, r, j, s[k])
            ws.cell(row=r, column=11, value=s["atipicos"]).border = BORDER
        r += 1
    ws.column_dimensions['A'].width = 60
    _autosize(ws)
    _footer(ws, r + 1)

    if res.get("correlaciones"):
        ws2 = wb.create_sheet("Correlaciones")
        cs = res["correlaciones"]["cols"]; mm = res["correlaciones"]["matrix"]
        _header_row(ws2, [""] + cs, 1)
        for i, name in enumerate(cs):
            ws2.cell(row=i + 2, column=1, value=name).font = BOLD
            for j in range(len(cs)):
                _num(ws2, i + 2, j + 2, mm[i][j])
        _autosize(ws2)
    return _save(wb)


def build_econometria(res):
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo"
    r = _title(ws, f"PronoStat — Econometría (MCO · {res.get('form_label','')})",
               f"Dependiente: {res.get('yname','Y')} · n = {res['n']} · k = {res['k']} · gl = {res['gl']}"
               + (" · errores robustos (White HC1)" if res.get("robust") else ""))
    ws.cell(row=r, column=1, value="Ecuación:").font = BOLD
    ws.cell(row=r, column=2, value=res["equation"]); r += 2

    r = _header_row(ws, ["Variable", "Coef. β", "Error est.", "t", "p-valor", "Sig.",
                         "β estand.", "IC inf.", "IC sup."], r)
    for c in res["coeficientes"]:
        ws.cell(row=r, column=1, value=c["var"]).border = BORDER
        _num(ws, r, 2, c["beta"]); _num(ws, r, 3, c["se"]); _num(ws, r, 4, c["t"])
        _num(ws, r, 5, c["p"])
        ws.cell(row=r, column=6, value=c["stars"]).border = BORDER
        _num(ws, r, 7, c["beta_std"] if c["beta_std"] is not None else None)
        _num(ws, r, 8, c["ic"][0]); _num(ws, r, 9, c["ic"][1])
        r += 1
    r += 1
    a = res["ajuste"]
    ws.cell(row=r, column=1, value="Bondad de ajuste").font = BOLD; r += 1
    r = _header_row(ws, ["Indicador", "Valor"], r)
    for k, v in [("R²", a["r2"]), ("R² ajustado", a["r2_adj"]), ("F global", a["F"]),
                 ("p (F)", a["pF"]), ("Error est. regresión", a["se_reg"]),
                 ("AIC", a["aic"]), ("BIC", a["bic"])]:
        ws.cell(row=r, column=1, value=k).border = BORDER
        _num(ws, r, 2, v); r += 1
    _autosize(ws)

    # Diagnósticos
    ws2 = wb.create_sheet("Diagnósticos")
    rr = _header_row(ws2, ["Prueba", "Estadístico", "p-valor / valor"], 1)
    d = res["diagnostics"]

    def add(name, stat, p):
        nonlocal rr
        ws2.cell(row=rr, column=1, value=name).border = BORDER
        c = ws2.cell(row=rr, column=2, value=stat); c.border = BORDER
        c2 = ws2.cell(row=rr, column=3, value=p); c2.border = BORDER
        if isinstance(p, (int, float)):
            c2.number_format = "0.0000"
        rr += 1
    if d.get("condition_number") == d.get("condition_number"):
        add("Número de condición", round(d["condition_number"], 2), "—")
    if d.get("durbin_watson") == d.get("durbin_watson"):
        add("Durbin-Watson", round(d["durbin_watson"], 4), "—")
    for key, nm in [("breusch_pagan", "Breusch-Pagan"), ("white", "White"),
                    ("breusch_godfrey", "Breusch-Godfrey"), ("jarque_bera", "Jarque-Bera"), ("reset", "Ramsey RESET")]:
        t = d.get(key)
        if t:
            stat = t.get("LM", t.get("JB", t.get("F")))
            add(nm, round(stat, 4) if stat is not None else "—", t.get("p"))
    if d.get("vif"):
        rr += 1
        ws2.cell(row=rr, column=1, value="VIF").font = BOLD; rr += 1
        rr = _header_row(ws2, ["Variable", "VIF"], rr)
        for v in d["vif"]:
            ws2.cell(row=rr, column=1, value=v["var"]).border = BORDER
            _num(ws2, rr, 2, v["vif"]); rr += 1
    _autosize(ws2)
    _interp_sheet(wb, res)
    return _save(wb)


def _interp_sheet(wb, res, title="Interpretación"):
    """Agrega una hoja con la interpretación educativa de los resultados."""
    lines = (res or {}).get("interpretacion") or []
    if not lines:
        return
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value="Interpretación de resultados").font = TITLE_FONT
    ws.cell(row=2, column=1, value="Lectura en lenguaje claro de los valores obtenidos.").font = SUB_FONT
    r = 4
    for line in lines:
        c = ws.cell(row=r, column=1, value="•  " + line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 34
        r += 1
    ws.column_dimensions['A'].width = 115
    _footer(ws, r)


# --------------------------------------------------------------------------- #
def build_dataset(columns, rows, source="dataset"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    r = _title(ws, "PronoStat — Conjunto de datos", f"Fuente: {source} · {len(rows)} filas · {len(columns)} columnas")
    names = [c["name"] for c in columns]
    r = _header_row(ws, names, r)
    for row in rows:
        for j, c in enumerate(names, 1):
            v = row.get(c)
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = BORDER
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0.####"
        r += 1
    _autosize(ws)
    _footer(ws, r + 1)
    return _save(wb)


def build_descriptiva(columna, res):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    r = _title(ws, f"PronoStat — Estadística descriptiva: {columna}", f"n = {res['n']}")
    s = res["summary"]
    labels = [
        ("Media", s["media"]), ("Mediana", s["mediana"]),
        ("Moda", ", ".join(f"{m:.4g}" for m in s["moda"]) if s["moda"] else "—"),
        ("Desviación estándar", s["desv_std"]), ("Varianza", s["varianza"]),
        ("Coef. de variación (%)", s["coef_var_pct"]), ("Error estándar", s["error_std"]),
        ("Mínimo", s["min"]), ("Q1", s["q1"]), ("Q2 (mediana)", s["q2"]), ("Q3", s["q3"]),
        ("Máximo", s["max"]), ("Rango", s["rango"]), ("IQR", s["iqr"]),
        ("Asimetría", s["asimetria"]), ("Curtosis", s["curtosis"]), ("Suma", s["suma"]),
        ("IC 95% media (inf.)", s["ic95_media"][0]), ("IC 95% media (sup.)", s["ic95_media"][1]),
        ("Valores atípicos", res["n_outliers"]),
    ]
    r = _header_row(ws, ["Estadístico", "Valor"], r)
    for k, v in labels:
        ws.cell(row=r, column=1, value=k).border = BORDER
        cell = ws.cell(row=r, column=2, value=v)
        cell.border = BORDER
        if isinstance(v, (int, float)):
            cell.number_format = "#,##0.0000"
        r += 1
    _autosize(ws)
    _footer(ws, r)

    # Hoja histograma
    ws2 = wb.create_sheet("Histograma")
    rr = _title(ws2, "Histograma", f"{columna}")
    rr = _header_row(ws2, ["Intervalo", "Marca de clase", "Frecuencia"], rr)
    h = res["hist"]
    for i in range(len(h["counts"])):
        ws2.cell(row=rr, column=1, value=h["labels"][i]).border = BORDER
        _num(ws2, rr, 2, h["centers"][i])
        ws2.cell(row=rr, column=3, value=h["counts"][i]).border = BORDER
        rr += 1
    _autosize(ws2)
    _interp_sheet(wb, res)
    return _save(wb)


def build_regresion(ctx, res):
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo"
    cy = ctx.get("cy", "Y"); cx = ctx.get("cx", "X")
    r = _title(ws, f"PronoStat — Regresión ({res['model']})", f"{cy} en función de {cx} · n = {res['n']}")
    ws.cell(row=r, column=1, value="Ecuación:").font = BOLD
    ws.cell(row=r, column=2, value=res["equation"])
    r += 2
    m = res["metrics"]
    r = _header_row(ws, ["Métrica", "Valor"], r)
    for k, v in [("R²", m["r2"]), ("R² ajustado", m["r2_adj"]), ("RMSE", m["rmse"]),
                 ("MAE", m["mae"]), ("SS residual", m["ss_res"]), ("SS total", m["ss_tot"])]:
        ws.cell(row=r, column=1, value=k).border = BORDER
        _num(ws, r, 2, v); r += 1
    r += 1
    ws.cell(row=r, column=1, value="Parámetros").font = BOLD; r += 1
    r = _header_row(ws, ["Parámetro", "Valor"], r)
    for k, v in res["params"].items():
        ws.cell(row=r, column=1, value=k).border = BORDER
        _num(ws, r, 2, v); r += 1
    _autosize(ws)

    # datos + ajuste + residuos
    ws2 = wb.create_sheet("Datos y residuos")
    rr = _header_row(ws2, [cx, cy, "Ajustado ŷ", "Residuo"], 1)
    px, py = res["points"]["x"], res["points"]["y"]
    for i in range(len(px)):
        _num(ws2, rr, 1, px[i]); _num(ws2, rr, 2, py[i])
        _num(ws2, rr, 3, res["fitted"][i]); _num(ws2, rr, 4, res["residuals"][i])
        rr += 1
    _autosize(ws2)

    if res.get("predictions"):
        ws3 = wb.create_sheet("Predicciones")
        rr = _header_row(ws3, ["X", "Ŷ"], 1)
        for p in res["predictions"]:
            _num(ws3, rr, 1, p["x"]); _num(ws3, rr, 2, p["y"]); rr += 1
        _autosize(ws3)
    _interp_sheet(wb, res)
    return _save(wb)


def build_pronostico(ctx, res):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    col = ctx.get("col", "serie")
    r = _title(ws, f"PronoStat — Pronóstico: {res['label']}", f"Serie: {col} · n = {res['n']} · horizonte = {res['h']}")

    ws.cell(row=r, column=1, value="Parámetros del modelo").font = BOLD; r += 1
    r = _header_row(ws, ["Parámetro", "Valor"], r)
    for k, v in res["params"].items():
        ws.cell(row=r, column=1, value=k).border = BORDER
        c = ws.cell(row=r, column=2, value=v); c.border = BORDER
        if isinstance(v, (int, float)):
            c.number_format = "#,##0.0000"
        r += 1
    ws.cell(row=r, column=1, value="σ residual").border = BORDER
    _num(ws, r, 2, res["sigma"]); r += 2

    ws.cell(row=r, column=1, value="Métricas de exactitud").font = BOLD; r += 1
    r = _header_row(ws, ["Métrica", "Ajuste (in-sample)", "Prueba (holdout)"], r)
    im = res["insample_metrics"]; tm = res.get("test_metrics")
    for k in ["MAE", "RMSE", "MAPE", "SMAPE", "MASE"]:
        ws.cell(row=r, column=1, value=k).border = BORDER
        _num(ws, r, 2, im[k])
        _num(ws, r, 3, tm[k] if tm else None)
        r += 1
    _autosize(ws)

    # Serie + ajuste
    ws2 = wb.create_sheet("Serie y ajuste")
    rr = _header_row(ws2, ["t", "Observado", "Ajustado"], 1)
    for i, v in enumerate(res["series"]):
        _num(ws2, rr, 1, i + 1); _num(ws2, rr, 2, v)
        _num(ws2, rr, 3, res["fitted"][i]); rr += 1
    _autosize(ws2)

    # Pronóstico + IC
    ws3 = wb.create_sheet("Pronóstico")
    conf = res["intervals"]["conf"]
    rr = _header_row(ws3, ["Paso", "Pronóstico", f"Límite inf. {conf}%", f"Límite sup. {conf}%"], 1)
    for i, v in enumerate(res["forecast"]):
        ws3.cell(row=rr, column=1, value=f"t+{i+1}").border = BORDER
        _num(ws3, rr, 2, v)
        _num(ws3, rr, 3, res["intervals"]["lower"][i])
        _num(ws3, rr, 4, res["intervals"]["upper"][i])
        rr += 1
    _autosize(ws3)

    if res.get("test"):
        ws4 = wb.create_sheet("Prueba")
        rr = _header_row(ws4, ["Real", "Predicho"], 1)
        t = res["test"]
        for i in range(len(t["actual"])):
            _num(ws4, rr, 1, t["actual"][i]); _num(ws4, rr, 2, t["pred"][i]); rr += 1
        _autosize(ws4)
    _interp_sheet(wb, res)
    return _save(wb)
